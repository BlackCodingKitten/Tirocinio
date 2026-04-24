from __future__ import annotations

import json
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

JsonDict = Dict[str, Any]

DEFAULT_CLASSIFICATION_JSON = Path("Data/TranscriptionData/final_classification/manual_revision.json")
STANDARD_OUTPUT_NAME = "3M_4M_merge.xlsx"
LOGPROB_OUTPUT_NAME = "1A_1B_logprob_merge.xlsx"
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBHEADER_FILL = PatternFill(fill_type="solid", fgColor="D9EAF7")
NUMBER_FORMAT = "0.0000"
PERCENT_FORMAT = "0.0000"


@dataclass(frozen=True)
class BinaryMetrics:
    """Support-weighted binary classification metrics derived from a confusion matrix."""

    accuracy: float
    precision: float
    recall: float
    f1_score: float
    support_0: int
    support_1: int
    valid_samples: int
    total_samples: int
    true_0_pred_0: int
    true_0_pred_1: int
    true_1_pred_0: int
    true_1_pred_1: int


@dataclass(frozen=True)
class StandardizedRow:
    """Single normalized output row ready to be written into the final workbook."""

    sheet_name: str
    payload: JsonDict


class CliError(ValueError):
    """Raised when the command line arguments are inconsistent."""


class WorkbookShapeError(ValueError):
    """Raised when an input workbook does not match any supported schema."""


def parse_cli(argv: Sequence[str]) -> Tuple[bool, Path, List[Path]]:
    """Parse CLI arguments.

    Supported usage examples:
        python report_merge.py 1A_evaluation_summary.xlsx 2_evaluation_summary.xlsx 3_evaluation_summary.xlsx 4_evaluation_summary.xlsx
        python report_merge.py logprob=TRUE 1A_evaluation_summary.xlsx 1B_evaluation_summary.xlsx
        python report_merge.py --classification-json Data/TranscriptionData/final_classification/final_results.json 1A_evaluation_summary.xlsx ...
    """

    logprob: bool = False
    classification_json: Path = DEFAULT_CLASSIFICATION_JSON
    file_paths: List[Path] = []

    i = 0
    while i < len(argv):
        token = argv[i]
        lowered = token.lower()

        if lowered == "logprob=true":
            logprob = True
            i += 1
            continue
        if lowered == "logprob=false":
            logprob = False
            i += 1
            continue
        if token == "--logprob":
            logprob = True
            i += 1
            continue
        if token == "--classification-json":
            if i + 1 >= len(argv):
                raise CliError("Missing path after --classification-json.")
            classification_json = Path(argv[i + 1])
            i += 2
            continue
        if lowered.startswith("classification_json="):
            classification_json = Path(token.split("=", 1)[1])
            i += 1
            continue

        file_paths.append(Path(token))
        i += 1

    if not file_paths:
        raise CliError("No input .xlsx files provided.")

    return logprob, classification_json, file_paths


def experiment_label_from_path(path: Path) -> str:
    """Return the experiment label as the filename prefix before the first underscore."""

    return path.stem.split("_", 1)[0]


def experiment_sort_key(label: str) -> Tuple[int, str]:
    """Create a stable sort key for labels such as 1A, 1B, 2, 3, 4."""

    number_part = ""
    suffix_part = ""
    for char in label:
        if char.isdigit() and suffix_part == "":
            number_part += char
        else:
            suffix_part += char
    number_value = int(number_part) if number_part else 10**9
    return number_value, suffix_part


def load_rows(ws: Worksheet) -> List[JsonDict]:
    """Read an Excel sheet into a list of dictionaries using the first row as header."""

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    data_rows: List[JsonDict] = []
    for values in rows[1:]:
        record: JsonDict = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            record[header] = values[index] if index < len(values) else None
        if any(value is not None for value in record.values()):
            data_rows.append(record)
    return data_rows


def get_first_present(row: Mapping[str, Any], keys: Sequence[str], default: Any = None) -> Any:
    """Return the first present key from a row."""

    for key in keys:
        if key in row and row[key] is not None:
            return row[key]
    return default


def safe_int(value: Any) -> int:
    """Convert an Excel value to int safely."""

    if value is None or value == "":
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return int(round(value))
    return int(str(value).strip())


def safe_float(value: Any) -> float:
    """Convert an Excel value to float safely."""

    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    return float(str(value).strip())


def divide(numerator: float, denominator: float) -> float:
    """Division guarded against division by zero."""

    return numerator / denominator if denominator else 0.0


def f1_score(precision: float, recall: float) -> float:
    """Compute the F1 score from precision and recall."""

    return divide(2.0 * precision * recall, precision + recall)


def build_binary_metrics(
    true_0_pred_0: int,
    true_0_pred_1: int,
    true_1_pred_0: int,
    true_1_pred_1: int,
    total_samples: Optional[int] = None,
) -> BinaryMetrics:
    """Compute support-weighted global metrics from binary confusion counts.

    Precision, recall and F1 are support-weighted across both labels so that the
    final report contains one single value per metric instead of one value per label.
    Accuracy is computed on total_samples when available; otherwise it falls back to
    the number of valid samples represented by the confusion counts.
    """

    support_0 = true_0_pred_0 + true_0_pred_1
    support_1 = true_1_pred_0 + true_1_pred_1
    valid_samples = support_0 + support_1
    resolved_total_samples = total_samples if total_samples is not None and total_samples >= valid_samples else valid_samples

    predicted_0 = true_0_pred_0 + true_1_pred_0
    predicted_1 = true_0_pred_1 + true_1_pred_1

    precision_0 = divide(true_0_pred_0, predicted_0)
    recall_0 = divide(true_0_pred_0, support_0)
    f1_0 = f1_score(precision_0, recall_0)

    precision_1 = divide(true_1_pred_1, predicted_1)
    recall_1 = divide(true_1_pred_1, support_1)
    f1_1 = f1_score(precision_1, recall_1)

    weighted_precision = divide((precision_0 * support_0) + (precision_1 * support_1), valid_samples)
    weighted_recall = divide((recall_0 * support_0) + (recall_1 * support_1), valid_samples)
    weighted_f1 = divide((f1_0 * support_0) + (f1_1 * support_1), valid_samples)
    accuracy = divide(true_0_pred_0 + true_1_pred_1, resolved_total_samples)

    return BinaryMetrics(
        accuracy=accuracy,
        precision=weighted_precision,
        recall=weighted_recall,
        f1_score=weighted_f1,
        support_0=support_0,
        support_1=support_1,
        valid_samples=valid_samples,
        total_samples=resolved_total_samples,
        true_0_pred_0=true_0_pred_0,
        true_0_pred_1=true_0_pred_1,
        true_1_pred_0=true_1_pred_0,
        true_1_pred_1=true_1_pred_1,
    )


def find_total_samples(row: Mapping[str, Any]) -> Optional[int]:
    """Return the total sample count using the supported schemas."""

    value = get_first_present(
        row,
        [
            "sample_count",
            "total_examples",
            "n_samples",
            "total_examples_count",
        ],
    )
    return safe_int(value) if value is not None else None


def extract_standard_confusion(row: Mapping[str, Any]) -> Tuple[int, int, int, int]:
    """Extract the binary confusion counts from a standard row.

    Returns:
        (true_0_pred_0, true_0_pred_1, true_1_pred_0, true_1_pred_1)
    """

    if all(key in row for key in ("true_negative", "false_positive", "false_negative", "true_positive")):
        return (
            safe_int(row["true_negative"]),
            safe_int(row["false_positive"]),
            safe_int(row["false_negative"]),
            safe_int(row["true_positive"]),
        )

    candidates = [
        ("true_0_pred_0", "true_0_pred_1", "true_1_pred_0", "true_1_pred_1"),
        ("free_actual_0_pred_0", "free_actual_0_pred_1", "free_actual_1_pred_0", "free_actual_1_pred_1"),
        ("score_actual_0_pred_0", "score_actual_0_pred_1", "score_actual_1_pred_0", "score_actual_1_pred_1"),
    ]

    for keys in candidates:
        if all(key in row for key in keys):
            return tuple(safe_int(row[key]) for key in keys)  # type: ignore[return-value]

    raise WorkbookShapeError(f"Unable to extract confusion counts from row with keys: {sorted(row.keys())}")


def extract_logprob_confusion(row: Mapping[str, Any], prediction_mode: str) -> Tuple[int, int, int, int]:
    """Extract the confusion counts for the 1B workbook in free/score mode."""

    if prediction_mode == "free":
        keys = ("free_actual_0_pred_0", "free_actual_0_pred_1", "free_actual_1_pred_0", "free_actual_1_pred_1")
    elif prediction_mode == "score":
        keys = ("score_actual_0_pred_0", "score_actual_0_pred_1", "score_actual_1_pred_0", "score_actual_1_pred_1")
    else:
        raise WorkbookShapeError(f"Unsupported prediction_mode={prediction_mode!r}")

    if not all(key in row for key in keys):
        raise WorkbookShapeError(
            f"Unable to extract {prediction_mode} confusion counts from row with keys: {sorted(row.keys())}"
        )
    return tuple(safe_int(row[key]) for key in keys)  # type: ignore[return-value]


def is_logprob_workbook(path: Path, workbook_sheet_names: Sequence[str]) -> bool:
    """Detect the special 1B workbook."""

    return experiment_label_from_path(path).upper() == "1B" or "score_confusion" in workbook_sheet_names or "free_confusion" in workbook_sheet_names


def build_standard_rows(file_path: Path, classification_map: Mapping[str, str]) -> List[StandardizedRow]:
    """Build normalized rows for the standard merge output."""

    wb = load_workbook(file_path, data_only=True)
    experiment = experiment_label_from_path(file_path)

    if is_logprob_workbook(file_path, wb.sheetnames):
        raise WorkbookShapeError(
            f"{file_path.name} is a log-probability workbook. Use logprob=TRUE to process it."
        )

    output_rows: List[StandardizedRow] = []

    for sheet_name in ("global_metrics", "category_metrics", "video_metrics"):
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        for row in load_rows(ws):
            total_samples = find_total_samples(row)
            t00, t01, t10, t11 = extract_standard_confusion(row)
            metrics = build_binary_metrics(t00, t01, t10, t11, total_samples)

            payload: JsonDict = {
                "experiment": experiment,
                "n_samples": metrics.total_samples,
                "valid_samples": metrics.valid_samples,
                "accuracy": metrics.accuracy,
                "precision": metrics.precision,
                "recall": metrics.recall,
                "f1_score": metrics.f1_score,
                "true_0_pred_0": metrics.true_0_pred_0,
                "true_0_pred_1": metrics.true_0_pred_1,
                "true_1_pred_0": metrics.true_1_pred_0,
                "true_1_pred_1": metrics.true_1_pred_1,
            }

            if sheet_name == "global_metrics":
                payload = {"scope": "GLOBAL", **payload}
            elif sheet_name == "category_metrics":
                payload = {
                    "normalized_question_category": str(row.get("normalized_question_category", "")),
                    **payload,
                }
            elif sheet_name == "video_metrics":
                video_id = str(row.get("video_id", ""))
                payload = {
                    "video_id": video_id,
                    "classification": classification_map.get(f"{video_id}.mp4", ""),
                    **payload,
                }

            output_rows.append(StandardizedRow(sheet_name=sheet_name, payload=payload))

    return output_rows


def build_logprob_rows(file_path: Path, classification_map: Mapping[str, str]) -> List[StandardizedRow]:
    """Build normalized rows for the logprob merge output.

    For standard workbooks (for example 1A), only a single 'free' view is produced.
    For the 1B workbook, both 'free' and 'score' views are produced.
    """

    wb = load_workbook(file_path, data_only=True)
    experiment = experiment_label_from_path(file_path)
    workbook_is_logprob = is_logprob_workbook(file_path, wb.sheetnames)

    output_rows: List[StandardizedRow] = []

    for sheet_name in ("global_metrics", "category_metrics", "video_metrics"):
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        rows = load_rows(ws)
        for row in rows:
            total_samples = find_total_samples(row)

            if workbook_is_logprob:
                modes = ("free", "score")
            else:
                modes = ("free",)

            for prediction_mode in modes:
                if workbook_is_logprob:
                    t00, t01, t10, t11 = extract_logprob_confusion(row, prediction_mode)
                else:
                    t00, t01, t10, t11 = extract_standard_confusion(row)

                metrics = build_binary_metrics(t00, t01, t10, t11, total_samples)

                payload: JsonDict = {
                    "experiment": experiment,
                    "prediction_mode": prediction_mode,
                    "n_samples": metrics.total_samples,
                    "valid_samples": metrics.valid_samples,
                    "accuracy": metrics.accuracy,
                    "precision": metrics.precision,
                    "recall": metrics.recall,
                    "f1_score": metrics.f1_score,
                    "true_0_pred_0": metrics.true_0_pred_0,
                    "true_0_pred_1": metrics.true_0_pred_1,
                    "true_1_pred_0": metrics.true_1_pred_0,
                    "true_1_pred_1": metrics.true_1_pred_1,
                    "avg_choice_0_score": safe_float(row.get("avg_choice_0_score")),
                    "avg_choice_1_score": safe_float(row.get("avg_choice_1_score")),
                    "avg_confidence": safe_float(row.get("avg_confidence")),
                    "avg_score_gap": safe_float(row.get("avg_score_gap")),
                    "free_vs_score_agreement_rate": safe_float(row.get("free_vs_score_agreement_rate")),
                }

                if sheet_name == "global_metrics":
                    payload = {"scope": "GLOBAL", **payload}
                elif sheet_name == "category_metrics":
                    payload = {
                        "normalized_question_category": str(row.get("normalized_question_category", "")),
                        **payload,
                    }
                elif sheet_name == "video_metrics":
                    video_id = str(row.get("video_id", ""))
                    payload = {
                        "video_id": video_id,
                        "classification": classification_map.get(f"{video_id}.mp4", ""),
                        **payload,
                    }

                output_rows.append(StandardizedRow(sheet_name=sheet_name, payload=payload))

    return output_rows


def load_classification_map(classification_json: Path) -> Dict[str, str]:
    """Load the external classification mapping used for the merged video report."""

    if not classification_json.exists():
        print(
            f"[WARNING] Classification JSON not found: {classification_json}. The 'classification' column will be left empty.",
            file=sys.stderr,
        )
        return {}

    with classification_json.open("r", encoding="utf-8") as handle:
        raw = json.load(handle)

    classification_map: Dict[str, str] = {}
    for key, value in raw.items():
        if not isinstance(value, dict):
            continue
        classification_map[str(key)] = str(value.get("classification", ""))
    return classification_map


def group_rows_by_sheet(rows: Iterable[StandardizedRow]) -> Dict[str, List[JsonDict]]:
    """Group standardized rows by target sheet name."""

    grouped: Dict[str, List[JsonDict]] = {
        "global_metrics": [],
        "category_metrics": [],
        "video_metrics": [],
    }
    for row in rows:
        grouped.setdefault(row.sheet_name, []).append(row.payload)
    return grouped


def sort_grouped_rows(grouped_rows: Dict[str, List[JsonDict]]) -> None:
    """Sort grouped rows in-place for a deterministic output."""

    grouped_rows["global_metrics"].sort(key=lambda item: (experiment_sort_key(str(item["experiment"])), str(item.get("prediction_mode", ""))))
    grouped_rows["category_metrics"].sort(
        key=lambda item: (
            str(item.get("normalized_question_category", "")),
            experiment_sort_key(str(item["experiment"])),
            str(item.get("prediction_mode", "")),
        )
    )
    grouped_rows["video_metrics"].sort(
        key=lambda item: (
            _video_sort_key(str(item.get("video_id", ""))),
            experiment_sort_key(str(item["experiment"])),
            str(item.get("prediction_mode", "")),
        )
    )


def _video_sort_key(video_id: str) -> Tuple[int, str]:
    """Sort video identifiers such as Video1, Video2, ..., Video100 naturally."""

    prefix = ""
    number = ""
    for char in video_id:
        if char.isdigit():
            number += char
        else:
            prefix += char
    return (int(number) if number else 10**9, prefix)


def write_table(ws: Worksheet, rows: List[JsonDict]) -> None:
    """Write a list of dictionaries to an Excel sheet with a clean report style."""

    if not rows:
        ws.append(["No data"])
        return

    headers = list(rows[0].keys())
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    float_like_headers = {
        "accuracy",
        "precision",
        "recall",
        "f1_score",
        "avg_choice_0_score",
        "avg_choice_1_score",
        "avg_confidence",
        "avg_score_gap",
        "free_vs_score_agreement_rate",
    }
    integer_like_headers = {
        "n_samples",
        "valid_samples",
        "support_0",
        "support_1",
        "true_0_pred_0",
        "true_0_pred_1",
        "true_1_pred_0",
        "true_1_pred_1",
    }

    for col_idx, header in enumerate(headers, start=1):
        column_letter = get_column_letter(col_idx)
        max_len = len(str(header))
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell_value = cell.value
            if cell_value is not None:
                max_len = max(max_len, len(str(cell_value)))

            if header in float_like_headers and isinstance(cell_value, (int, float)):
                cell.number_format = NUMBER_FORMAT
            elif header in integer_like_headers and isinstance(cell_value, (int, float)):
                cell.number_format = "0"

        ws.column_dimensions[column_letter].width = min(max_len + 2, 28)

    for row_idx in range(2, ws.max_row + 1):
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = SUBHEADER_FILL


def build_output_workbook(grouped_rows: Dict[str, List[JsonDict]], output_path: Path) -> None:
    """Create the final merged workbook."""

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    for sheet_name in ("global_metrics", "category_metrics", "video_metrics"):
        ws = wb.create_sheet(title=sheet_name)
        write_table(ws, grouped_rows.get(sheet_name, []))

    wb.save(output_path)


def validate_standard_mode(file_paths: Sequence[Path]) -> None:
    """Validate the allowed input files for the standard merge."""

    for path in file_paths:
        if experiment_label_from_path(path).upper() == "1B":
            raise CliError(
                "The file 1B_evaluation_summary.xlsx must be processed with logprob=TRUE because it contains both free and score metrics."
            )


def validate_logprob_mode(file_paths: Sequence[Path]) -> None:
    """Validate the allowed input files for the logprob merge."""

    allowed = {"1A", "1B"}
    labels = {experiment_label_from_path(path).upper() for path in file_paths}
    unsupported = sorted(label for label in labels if label not in allowed)
    if unsupported:
        raise CliError(
            "logprob=TRUE is reserved for the experiment 1 comparison. Unsupported labels found: "
            + ", ".join(unsupported)
        )
    if "1B" not in labels:
        raise CliError("logprob=TRUE requires at least 1B_evaluation_summary.xlsx in input.")


def ensure_input_files_exist(file_paths: Sequence[Path]) -> None:
    """Fail early if an input file does not exist."""

    missing = [str(path) for path in file_paths if not path.exists()]
    if missing:
        raise CliError("Missing input files: " + ", ".join(missing))


def main(argv: Sequence[str]) -> int:
    """Entrypoint used by the command line interface."""

    try:
        logprob, classification_json, file_paths = parse_cli(argv)
        ensure_input_files_exist(file_paths)
        classification_map = load_classification_map(classification_json)

        if logprob:
            validate_logprob_mode(file_paths)
            rows: List[StandardizedRow] = []
            for file_path in file_paths:
                rows.extend(build_logprob_rows(file_path, classification_map))
            output_path = Path(LOGPROB_OUTPUT_NAME)
        else:
            validate_standard_mode(file_paths)
            rows = []
            for file_path in file_paths:
                rows.extend(build_standard_rows(file_path, classification_map))
            output_path = Path(STANDARD_OUTPUT_NAME)

        grouped_rows = group_rows_by_sheet(rows)
        sort_grouped_rows(grouped_rows)
        build_output_workbook(grouped_rows, output_path)

        print(f"Created {output_path}")
        return 0

    except (CliError, WorkbookShapeError, OSError, json.JSONDecodeError) as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
